# -*- coding:utf-8 -*-

from openpyxl import (
    Workbook,
    load_workbook
)
import logging
import os
from data.models import Materials, Price, Brands
from data.database import db
import json

logger = logging.getLogger(__name__)

class SaveData:
    __file_name__ = r'finally/save_data.xlsx'

    def __init__(self, file_name = r'finally/save_data.xlsx'):
        __file_name__ = file_name
        if not os.path.isfile(__file_name__):
            wb = Workbook()
            wb.save(__file_name__)

    def save(self, name, data):
        wb = load_workbook(self.__file_name__)
        ws = wb.active
        list = []
        list.append(name.encode('utf-8'))
        for key, value in data.items():
            list.append(key.encode('utf-8'))
            list.append(value.encode('utf-8'))
        ws.append(list)
        wb.save(self.__file_name__)

    def save_brand_to_database(self, brand):
        tmp = db.session.query(Brands).filter_by(name=brand.get('name', 'None')).first()
        if tmp is not None:
            logger.error(tmp)
            return
        brands = Brands(
            name=brand.get('name', 'None'),
            url=brand.get('url', 'None'),
            desc=brand.get('desc', 'None')
        )
        brands.save()

    def save_material_to_database(self, materials):
        mater = db.session.query(Materials).filter_by(number=materials.get('number', 'None')).first()
        if mater is not None:
            logger.error(mater)
            return
        brand_name = materials.get('brand', 'None')
        brand = db.session.query(Brands).filter(Brands.name.like('%' + brand_name + '%')).first()
        if brand is None:
            logger.error(brand_name + " don't exist!" )
            return
        material = Materials(
            name=materials.get('name', 'None'),
            category=materials.get('category', 'None'),
            model=materials.get('model', 'None'),
            number=materials.get('number', 'None'),
            package=materials.get('package', 'None')
        )
        price = Price(price=json.dumps(materials.get('price', {'none':'none'})))
        material.price.append(price)
        material.brands = brand
        material.save()
