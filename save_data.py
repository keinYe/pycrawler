# -*- coding:utf-8 -*-

from openpyxl import (
    Workbook,
    load_workbook
)
import os

class SaveData:
    __file_name__ = r'finally/save_data.xlsx'

    def __init__(self, file_name = r'finally/save_data.xlsx'):
        __file_name__ = file_name
        if not os.path.isfile(__file_name__):
            wb = Workbook()
            wb.save(__file_name__)

    def save(self, name, data):
        wb = load_workbook(self.__file_name__)
        ws = wb.active
        list = []
        list.append(name.encode('utf-8'))
        # print(data)
        for n in data:
            # print(n)
            list.append(n[0].encode('utf-8'))
            list.append(n[1].encode('utf-8'))
        ws.append(list)
        wb.save(self.__file_name__)
